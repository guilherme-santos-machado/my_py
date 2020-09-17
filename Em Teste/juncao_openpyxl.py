import openpyxl as xl

cont = 0
# Abre o arquivo de origem 
filename ="./ATU_ESCOLAS_2018.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0] 
  
# Abre o arquivo de destino  
filename1 ="./dani.xlsx"
wb2 = xl.load_workbook(filename1, data_only = 'True') 
ws2 = wb2.active 
  
# Pega número de linhas e colunas 
mr = ws1.max_row 
mc = ws1.max_column 
linha = ws2.max_row + 1
  
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
    print('trabalhando linha ',i)
    for j in range (1, mc + 1):         
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
        # writing the read value to destination excel file 
        ws2.cell(row = i + linha, column = j).value = c.value 
    cont += 1
    if cont == 40000:
        print('salvando linhas')
        wb2.save(str(filename1)) 
        print('carregando planilha')
        wb2 = xl.load_workbook(filename1)
        print('ativando planilha') 
        ws2 = wb2.active 
        cont = 0
# saving the destination excel file 
wb2.save(str(filename1)) 
print('fim')