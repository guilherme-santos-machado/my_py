import openpyxl as xl

cont = 0
# Abre o arquivo de origem 
filename ="./teste.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0] 
ws1 = wb1.active 

mc = ws1.max_column 
i = 6
fim = 9


for j in range (1, mc + 1):         
        # reading cell value from source excel file 
        c = ws1.cell(row = i, column = j) 
        # writing the read value to destination excel file 
        ws1.cell(row = fim, column = j).value = c.value 


print('deletando linha')
ws1.delete_rows()